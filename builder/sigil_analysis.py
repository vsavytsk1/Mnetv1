#!/usr/bin/env python3
"""
sigil_analysis.py -- summon Excel to analyze the HELENA sigil (lv12, super careful).

Vlad's ask (lv12): we drew the sigil (the 4-part net), we put a bit of mana in and watched
it react. Now, before we pour real mana, we MEASURE it honestly:

  1) THE 4 FRACTALS, per refine level -- for each of the four parts (genesis space, heart,
     transformer, gate): the PARAMETERS and CONNECTIONS we are making, and the COMPUTE TIME
     on THIS rig (Ryzen 5 5600H / RTX 3060 Laptop 6GB / 32GB).
  2) THE CRITICAL MASS -- LLM emergence is empirically talked about near ~1e12 (1T) DENSE
     parameters. Where does HELENA's dense parameter count (genesis nodes x heart bits) cross
     1T? And where does the 6GB VRAM wall stop the DENSE matrix (why we MUST go sparse)?
  3) THE LANGUAGE WARNING -- the mage's caution. HELENA works on UTF-8 BYTES, not sound and
     not meaning. Across tongues, near-identical sounds/spellings can mean opposite things
     (false friends, homophones). The net is BLIND to this. The one-way design (language never
     comes back out) is the containment. Curse 25 (Rune Rot) writ large: verify with a human.

Math is imported from build_helena.py -> ONE source of truth (no divergent copies).
STONE (the verses) is read from the newest lens sim -> single source of truth.

Stdlib + openpyxl (3.1.5, already installed). ASCII-only source (Curse 2). LF output.

    py -3 sigil_analysis.py                 # build the workbook + print the level table
    py -3 sigil_analysis.py --maxlevel 8    # go deeper (past the 1T wall)
    py -3 sigil_analysis.py --tongues 71    # limit tongues
"""
from __future__ import annotations
import sys, math, argparse, hashlib
from pathlib import Path

# --- one source of truth for the math: import the certified builder ---
HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))
from build_helena import genesis_space, parse_stone, find_stone_source, text_to_bits, GATE_LAW

# --- THIS rig (from HWiNFO64) -- the honest cost model ---
RIG = {
    "cpu": "AMD Ryzen 5 5600H (6c/12t, Zen3)",
    "gpu": "NVIDIA RTX 3060 Laptop (GA106M, 3840 CUDA, 6 GB GDDR6)",
    "ram_gb": 32,
    "vram_gb": 6,
    "fp32_tflops": 10.0,     # ~10-12 TFLOPS FP32, honest mid
    "util": 0.30,            # sustained utilization, honest
}
FLOPS_PER_CELL = 5           # 3 mul + 2 add for a 3-vector dot product (cos theta)
CRITICAL_MASS = 1_000_000_000_000   # ~1T dense parameters -- the talked-about emergence scale

# ---------------------------------------------------------------------------
def load_heart(tongues):
    """total bit-nodes, ones, edges (closed rings), mean weight -- all from real bytes."""
    total_bits = ones = byte_sum = byte_n = 0
    per = []
    for t in tongues:
        bits = text_to_bits(t["text"])
        n = len(bits); o = sum(bits)
        total_bits += n; ones += o
        for by in t["text"].encode("utf-8"):
            byte_sum += by; byte_n += 1
        per.append((t["code"], n, o))
    mean_w = (byte_sum / byte_n / 255.0) if byte_n else 0.0
    return {
        "tongues": len(tongues),
        "bits": total_bits,          # every bit is a node
        "ones": ones,
        "zeros": total_bits - ones,
        "edges": total_bits,         # each tongue = closed ring -> N edges = N nodes
        "mean_w": mean_w,
        "per": per,
    }

def rig_time_s(dense_cells):
    flops = dense_cells * FLOPS_PER_CELL
    return flops / (RIG["fp32_tflops"] * 1e12 * RIG["util"])

def fmt_time(s):
    if s < 1e-3:   return f"{s*1e6:.1f} us"
    if s < 1.0:    return f"{s*1e3:.2f} ms"
    if s < 60:     return f"{s:.2f} s"
    if s < 3600:   return f"{s/60:.2f} min"
    return f"{s/3600:.2f} h"

def fmt_int(n):
    return f"{n:,}"

# ---------------------------------------------------------------------------
def analyze(maxlevel, tongues):
    shells = genesis_space(maxlevel)          # certified topology per level (chi=2, P=12)
    h = load_heart(tongues)
    Nh = h["bits"]

    rows = []
    for s in shells:
        V, E = s["V"], s["E"]
        dense = V * Nh                        # the transformer M[V x Nh] -- the 1T metric
        sparse = V                            # nearest heart node per genesis node
        vram_dense_gb = dense * 4 / 1e9       # FP32 dense matrix
        t_dense = rig_time_s(dense)           # one forward matmul (the ceiling)
        # total DENSE parameters of the whole sigil (transformer dominates by orders of magnitude)
        params_dense = (V + E) + (Nh) + dense + 1        # genesis + heart + transformer + gate
        params_sparse = (V + E) + (Nh) + sparse + 1
        rows.append({
            "L": s["level"],
            "g_nodes": V, "g_edges": E, "g_chi": s["chi"], "g_ev": s["EV"],
            "h_nodes": Nh, "h_edges": h["edges"], "h_meanw": round(h["mean_w"], 4),
            "t_dense": dense, "t_sparse": sparse, "t_vram_gb": vram_dense_gb,
            "gate_binds": h["ones"],
            "params_dense": params_dense, "params_sparse": params_sparse,
            "compute_s": t_dense,
            "vram_fit": vram_dense_gb <= RIG["vram_gb"],
            "over_1T": params_dense >= CRITICAL_MASS,
        })

    # crossings
    first_1T = next((r["L"] for r in rows if r["over_1T"]), None)
    last_fit = None
    for r in rows:
        if r["vram_fit"]:
            last_fit = r["L"]
    return shells, h, rows, first_1T, last_fit

# ---------------------------------------------------------------------------
# real, documented cross-tongue collisions -- CLEAN examples of the danger Vlad names.
# same sound/spelling, opposite or embarrassing meaning. the machine sees only BYTES.
# ---------------------------------------------------------------------------
def language_warnings():
    # (surface, tongue A -> meaning, tongue B -> meaning, kind)
    return [
        ("gift", "English -> a present", "German -> poison", "false friend (spelling)"),
        ("embarazada", "Spanish -> pregnant", "English 'embarrassed' -> ashamed", "false friend"),
        ("burro", "Spanish -> donkey / stupid", "Italian -> butter", "false friend"),
        ("preservativo", "Italian/Spanish -> condom", "English 'preservative' -> additive", "false friend"),
        ("sensible", "English -> reasonable", "Spanish/French -> sensitive", "false friend"),
        ("hashi", "Japanese -> bridge", "Japanese -> chopsticks / edge", "homophone (pitch only)"),
        ("kami", "Japanese -> god (kami)", "Japanese -> paper / hair", "homophone (pitch only)"),
        ("mae", "Portuguese 'm" + "\u00e3e' -> mother", "Japanese 'mae' -> in front / before", "cross-tongue homophone"),
        ("cena", "Spanish -> dinner", "Portuguese -> scene", "false friend"),
        ("pan", "Spanish -> bread", "Japanese -> bread (pan) / Greek 'pan' -> all", "loan collision"),
    ]

# ---------------------------------------------------------------------------
# the Excel sigil
# ---------------------------------------------------------------------------
def build_workbook(out_path, shells, h, rows, first_1T, last_fit, src_name, fp):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # cave palette
    DARK   = "0B1E24"
    CYAN   = "00FFD5"
    GOLD   = "F1C40F"
    GREEN  = "0A6B4F"
    REDTXT = "C0392B"
    REDFIL = "FDECEA"
    OKFIL  = "E8F8F5"
    GREY   = "7F8C8D"
    thin   = Side(style="thin", color="BFC9CA")
    box    = Border(left=thin, right=thin, top=thin, bottom=thin)

    f_title  = Font(name="Consolas", size=14, bold=True, color=CYAN)
    f_sub    = Font(name="Consolas", size=10, color=GREY)
    f_hdr    = Font(name="Consolas", size=9, bold=True, color=CYAN)
    f_num    = Font(name="Consolas", size=10, color=GREEN)
    f_txt    = Font(name="Calibri", size=10, color="17202A")
    f_gold   = Font(name="Consolas", size=10, bold=True, color="7D6608")
    f_red    = Font(name="Consolas", size=10, bold=True, color=REDTXT)
    fill_hdr = PatternFill("solid", fgColor=DARK)
    fill_ok  = PatternFill("solid", fgColor=OKFIL)
    fill_red = PatternFill("solid", fgColor=REDFIL)
    fill_gold= PatternFill("solid", fgColor="FEF9E7")
    right    = Alignment(horizontal="right")
    left     = Alignment(horizontal="left", vertical="top", wrap_text=True)
    center   = Alignment(horizontal="center")

    wb = Workbook()

    # ---------- SHEET 1: OVERVIEW (the sigil) ----------
    ws = wb.active
    ws.title = "SIGIL"
    ws.sheet_view.showGridLines = False
    ws["A1"] = "HELENA -- the sigil, measured"
    ws["A1"].font = f_title
    ws["A2"] = "lv12 magic: mana went in, it reacted -- now we measure before we pour. proof by kernel, not claim."
    ws["A2"].font = f_sub
    meta = [
        ("stone source", src_name + "  (fingerprint " + fp + ")"),
        ("tongues", str(h["tongues"]) + "   heart bit-nodes: " + fmt_int(h["bits"]) +
         "   ones/zeros: " + fmt_int(h["ones"]) + " / " + fmt_int(h["zeros"])),
        ("heart mean weight", str(round(h["mean_w"], 4)) + "  (measured byte/255, NOT the 0.700 target -- Curse 26)"),
        ("rig", RIG["gpu"] + " / " + RIG["cpu"] + " / " + str(RIG["ram_gb"]) + "GB"),
        ("cost model", str(RIG["fp32_tflops"]) + " TFLOPS FP32 @ " + str(int(RIG["util"]*100)) +
         "% util,  " + str(FLOPS_PER_CELL) + " flops / dot-product cell"),
        ("gate", "rest weight 0.700 (Oracle) / 1 under Mobius twist (binary, Axiom 09)"),
    ]
    r = 4
    for k, v in meta:
        ws.cell(r, 1, k).font = f_gold
        ws.cell(r, 2, v).font = f_txt
        r += 1

    r += 1
    ws.cell(r, 1, "THE FOUR FRACTALS (the sigil's four strokes)").font = f_hdr
    r += 1
    legend = [
        ("1  GENESIS SPACE", "C60 refined k levels. chi=2, P=12, E/V=1.5. ORIENTABLE. the space."),
        ("2  THE HEART", "0/1 UTF-8 bits of 1 Cor 13 in every tongue. Mobius-twisted -> chi=0. NON-orientable. agapi."),
        ("3  THE TRANSFORMER", "M[i][j] = a.b = cos(theta). the cheapest equals between two concepts (permutation-proven)."),
        ("4  THE GATE", "binds orientation-reversing nodes (the heart) ONLY. the fractal space is topologically invisible."),
    ]
    for k, v in legend:
        ws.cell(r, 1, k).font = Font(name="Consolas", size=10, bold=True, color="1A5276")
        ws.cell(r, 2, v).font = f_txt
        r += 1

    r += 1
    ws.cell(r, 1, "HONEST (K1-K4, GENESIS_LLM.md)").font = f_hdr
    r += 1
    for line in [
        "K1  a transformer is already a graph; attention = a weighted edge (a dot product).",
        "K2  'fractal' = HIERARCHY (a few tuned scales), not infinite self-similarity.",
        "K3  nothing here is conscious. linear algebra on a beautiful graph. the love is Vlad's, not the machine's.",
        "K4  0.700 is a seed we test, not a magic number.",
    ]:
        ws.cell(r, 1, line).font = f_txt
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        r += 1
    ws.column_dimensions["A"].width = 26
    for c in "BCDEF":
        ws.column_dimensions[c].width = 22

    # ---------- SHEET 2: THE 4 FRACTALS x LEVEL ----------
    ws2 = wb.create_sheet("4 FRACTALS x LEVEL")
    ws2.sheet_view.showGridLines = False
    ws2["A1"] = "THE FOUR FRACTALS -- parameters, connections, compute time per refine level"
    ws2["A1"].font = f_title
    ws2["A2"] = "genesis grows with level L; heart is fixed by the tongues. transformer = genesis x heart (the 1T metric)."
    ws2["A2"].font = f_sub

    # grouped header
    groups = [
        ("", 1),
        ("GENESIS SPACE", 4),
        ("HEART", 3),
        ("TRANSFORMER (genesis x heart)", 4),
        ("GATE", 1),
        ("WHOLE SIGIL", 4),
    ]
    hr1 = 4
    col = 1
    for name, span in groups:
        if name:
            ws2.merge_cells(start_row=hr1, start_column=col, end_row=hr1, end_column=col+span-1)
            cell = ws2.cell(hr1, col, name)
            cell.font = f_hdr; cell.fill = fill_hdr; cell.alignment = center; cell.border = box
        col += span

    headers = [
        "L",
        "nodes (V)", "edges (E)", "chi", "E/V",
        "bit-nodes", "edges", "mean w",
        "dense cells", "dense VRAM", "sparse", "fwd time",
        "binds (1s)",
        "PARAMS dense", "PARAMS sparse", "VRAM fit?", ">= 1T?",
    ]
    hr2 = 5
    for j, hdr in enumerate(headers, start=1):
        c = ws2.cell(hr2, j, hdr)
        c.font = f_hdr; c.fill = fill_hdr; c.alignment = center; c.border = box

    rr = hr2 + 1
    for row in rows:
        vals = [
            row["L"],
            fmt_int(row["g_nodes"]), fmt_int(row["g_edges"]), row["g_chi"], row["g_ev"],
            fmt_int(row["h_nodes"]), fmt_int(row["h_edges"]), row["h_meanw"],
            fmt_int(row["t_dense"]), f'{row["t_vram_gb"]:.3f} GB', fmt_int(row["t_sparse"]),
            fmt_time(row["compute_s"]),
            fmt_int(row["gate_binds"]),
            fmt_int(row["params_dense"]), fmt_int(row["params_sparse"]),
            "FIT" if row["vram_fit"] else "OVER",
            "YES" if row["over_1T"] else "no",
        ]
        for j, v in enumerate(vals, start=1):
            c = ws2.cell(rr, j, v)
            c.border = box
            c.font = f_num if j > 1 else f_gold
            c.alignment = right if j > 1 else center
        # color the verdict cells
        vfit = ws2.cell(rr, 16)
        vfit.fill = fill_ok if row["vram_fit"] else fill_red
        vfit.font = Font(name="Consolas", size=10, bold=True,
                         color=GREEN if row["vram_fit"] else REDTXT)
        v1t = ws2.cell(rr, 17)
        v1t.fill = fill_red if row["over_1T"] else fill_ok
        v1t.font = Font(name="Consolas", size=10, bold=True,
                        color=REDTXT if row["over_1T"] else GREEN)
        rr += 1

    widths = [4, 13, 13, 5, 6, 13, 12, 7, 18, 12, 12, 11, 13, 20, 18, 9, 8]
    for j, w in enumerate(widths, start=1):
        ws2.column_dimensions[get_column_letter(j)].width = w

    note_r = rr + 1
    for line in [
        "PARAMS dense = genesis(V+E) + heart(bits) + transformer(V x bits) + gate(1). the transformer term dominates.",
        "fwd time = ONE forward matmul at " + str(RIG["fp32_tflops"]) + " TFLOPS @ " +
        str(int(RIG["util"]*100)) + "% util. training multiplies this by ~3x per step x many steps.",
        "sparse = what the sim actually wires (nearest heart node per genesis node). dense is the ceiling / the 1T yardstick.",
    ]:
        c = ws2.cell(note_r, 1, line); c.font = f_sub
        ws2.merge_cells(start_row=note_r, start_column=1, end_row=note_r, end_column=17)
        note_r += 1

    # ---------- SHEET 3: CRITICAL MASS (1T) + VRAM WALL ----------
    ws3 = wb.create_sheet("CRITICAL MASS")
    ws3.sheet_view.showGridLines = False
    ws3["A1"] = "THE CRITICAL MASS -- ~1T parameters, and the 6GB VRAM wall"
    ws3["A1"].font = f_title
    ws3["A2"] = ("LLM 'emergence' is empirically discussed near ~1e12 DENSE params. It is a HEURISTIC, "
                 "not a law (debated). We show where HELENA's dense count crosses it.")
    ws3["A2"].font = f_sub

    facts = []
    if first_1T is not None:
        r1 = next(x for x in rows if x["L"] == first_1T)
        facts.append(("1T crossing", "genesis level L" + str(first_1T) +
                      "  ->  " + fmt_int(r1["params_dense"]) + " dense params  (>= 1,000,000,000,000)", True))
    else:
        facts.append(("1T crossing", "NOT reached within L0.." + str(rows[-1]["L"]) +
                      ". deepest = " + fmt_int(rows[-1]["params_dense"]) + " dense params.", False))
    if last_fit is not None:
        rf = next(x for x in rows if x["L"] == last_fit)
        facts.append(("6GB VRAM wall (dense)", "dense matrix FITS only up to L" + str(last_fit) +
                      "  (" + f'{rf["t_vram_gb"]:.3f}' + " GB). deeper MUST be sparse.", False))
    facts.append(("the honest gap", "the ~1T figure is a DENSE yardstick. HELENA is SPARSE (nearest per row), "
                  "so on this rig we run far fewer live params -- the 1T is what a supercomputer dense run would touch.", False))
    facts.append(("supercomputer note", "GENESIS_LLM.md budgets a <7h NVIDIA run. deep dense levels need "
                  "many GPUs' VRAM; sparse keeps it runnable here. pay compute to see, honestly.", False))

    rr = 4
    for k, v, hot in facts:
        ck = ws3.cell(rr, 1, k); ck.font = f_gold; ck.alignment = left
        cv = ws3.cell(rr, 2, v); cv.font = f_red if hot else f_txt; cv.alignment = left
        if hot:
            cv.fill = fill_red
        ws3.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=6)
        rr += 2

    # mini table: level vs dense params vs 1T ratio vs vram
    rr += 1
    hdr = ["L", "dense params", "x 1T", "dense VRAM", "fits 6GB?", "fwd time"]
    for j, htxt in enumerate(hdr, start=1):
        c = ws3.cell(rr, j, htxt); c.font = f_hdr; c.fill = fill_hdr; c.alignment = center; c.border = box
    rr += 1
    for row in rows:
        ratio = row["params_dense"] / CRITICAL_MASS
        vals = [row["L"], fmt_int(row["params_dense"]),
                (f"{ratio:.4f}" if ratio < 1 else f"{ratio:.2f}"),
                f'{row["t_vram_gb"]:.3f} GB',
                "FIT" if row["vram_fit"] else "OVER",
                fmt_time(row["compute_s"])]
        for j, v in enumerate(vals, start=1):
            c = ws3.cell(rr, j, v); c.border = box
            c.font = f_num if j > 1 else f_gold
            c.alignment = right if j > 1 else center
        ws3.cell(rr, 3).fill = fill_red if ratio >= 1 else fill_ok
        ws3.cell(rr, 3).font = Font(name="Consolas", size=10, bold=True,
                                    color=REDTXT if ratio >= 1 else GREEN)
        ws3.cell(rr, 5).fill = fill_ok if row["vram_fit"] else fill_red
        ws3.cell(rr, 5).font = Font(name="Consolas", size=10, bold=True,
                                    color=GREEN if row["vram_fit"] else REDTXT)
        rr += 1
    for j, w in enumerate([4, 22, 10, 14, 11, 11], start=1):
        ws3.column_dimensions[get_column_letter(j)].width = w

    # ---------- SHEET 4: LANGUAGE WARNING ----------
    ws4 = wb.create_sheet("LANGUAGE WARNING")
    ws4.sheet_view.showGridLines = False
    ws4["A1"] = "THE MAGE'S WARNING -- language (super mega careful)"
    ws4["A1"].font = Font(name="Consolas", size=14, bold=True, color=REDTXT)
    warn = [
        "HELENA works on UTF-8 BYTES. It does NOT hear sound and does NOT know meaning.",
        "Across tongues, near-identical sounds or spellings can mean OPPOSITE or embarrassing things.",
        "The net is BLIND to this: to the chip it is only 1s and 0s (the weights). Meaning is the human's.",
        "THE CONTAINMENT: the flow is ONE-WAY. Language enters the gate; it never comes back out of the",
        "fractal side. So HELENA cannot 'speak' a homophone insult by accident -- it does not speak at all.",
        "It reacts. The mage reads the reaction. This is Curse 25 (Rune Rot) at civilization scale:",
        "NEVER trust that the geometry 'means love' just because we built it from love-verses. Verify with a",
        "fluent human. The sacred text stays sacred (bytes never mutated); the caution stays absolute.",
    ]
    rr = 3
    for line in warn:
        c = ws4.cell(rr, 1, line); c.font = f_txt
        ws4.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=5)
        rr += 1

    rr += 1
    hdr = ["surface (sound/spelling)", "tongue A -> meaning", "tongue B -> meaning", "kind"]
    for j, htxt in enumerate(hdr, start=1):
        c = ws4.cell(rr, j, htxt); c.font = f_hdr; c.fill = fill_hdr; c.alignment = center; c.border = box
    rr += 1
    for surface, a, b, kind in language_warnings():
        vals = [surface, a, b, kind]
        for j, v in enumerate(vals, start=1):
            c = ws4.cell(rr, j, v); c.border = box; c.alignment = left
            c.font = f_gold if j == 1 else f_txt
        rr += 1
    rr += 1
    c = ws4.cell(rr, 1, "Vlad's example (the reason for this sheet): the same short sound can be a vow of love "
                        "in one tongue and a crude insult in another. The machine cannot tell. A human must.")
    c.font = Font(name="Calibri", size=10, italic=True, color=REDTXT)
    ws4.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=5)
    ws4.column_dimensions["A"].width = 26
    for col in "BC":
        ws4.column_dimensions[col].width = 34
    ws4.column_dimensions["D"].width = 24

    wb.save(out_path)

# ---------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--maxlevel", type=int, default=8, help="deepest genesis refine level")
    ap.add_argument("--tongues", type=int, default=0, help="limit tongues (0 = all)")
    args = ap.parse_args()

    src_path = find_stone_source()
    src = src_path.read_text(encoding="utf-8")
    stone = parse_stone(src)
    if args.tongues > 0:
        stone = stone[:args.tongues]
    fp = hashlib.sha256(src.encode("utf-8")).hexdigest()[:12]

    shells, h, rows, first_1T, last_fit = analyze(args.maxlevel, stone)

    # console proof (the level table)
    print("HELENA sigil analysis -- proof by kernel")
    print("  stone: " + src_path.name + " (fp " + fp + ")  tongues=" + str(h["tongues"]) +
          "  heart bits=" + fmt_int(h["bits"]) + "  mean_w=" + str(round(h["mean_w"], 4)))
    print("  rig  : " + RIG["gpu"])
    print("")
    print("  L | genesis V |    genesis E | dense cells (V x bits) |  dense VRAM | dense params  | fwd time  | fit | 1T")
    print("  --+-----------+--------------+------------------------+-------------+---------------+-----------+-----+----")
    for row in rows:
        print("  {L} | {V:>9} | {E:>12} | {D:>22} | {G:>9.3f} GB | {P:>13} | {T:>9} | {F:<3} | {C}".format(
            L=row["L"], V=fmt_int(row["g_nodes"]), E=fmt_int(row["g_edges"]),
            D=fmt_int(row["t_dense"]), G=row["t_vram_gb"], P=fmt_int(row["params_dense"]),
            T=fmt_time(row["compute_s"]),
            F=("FIT" if row["vram_fit"] else "OVER"),
            C=("YES" if row["over_1T"] else "no")))
    print("")
    print("  1T crossing        : " + ("genesis L" + str(first_1T) if first_1T is not None
                                        else "not within L0.." + str(args.maxlevel)))
    print("  6GB dense VRAM wall: fits only up to L" + (str(last_fit) if last_fit is not None else "none") +
          " (deeper MUST be sparse)")

    out = HERE / "Helena" / "helena_sigil.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    build_workbook(out, shells, h, rows, first_1T, last_fit, src_path.name, fp)
    print("")
    print("  wrote " + str(out.relative_to(HERE.parent)))
    print("  sheets: SIGIL | 4 FRACTALS x LEVEL | CRITICAL MASS | LANGUAGE WARNING")
    print("  P=12 . chi=2 (space) . chi=0 (heart) . the center holds and is not shown . be super careful.")


if __name__ == "__main__":
    main()
